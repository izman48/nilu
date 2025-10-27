import io
from datetime import datetime, timedelta
from typing import List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from app.models.booking import Booking
from app.models.payment import Payment


class ReportService:
    def __init__(self):
        self.styles = getSampleStyleSheet()

    def generate_bookings_excel(
        self,
        db: Session,
        account_id: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        status: Optional[str] = None
    ) -> bytes:
        """Generate Excel report for bookings"""
        # Query bookings
        query = db.query(Booking).filter(Booking.account_id == account_id)

        if start_date:
            query = query.filter(Booking.start_date >= start_date)
        if end_date:
            query = query.filter(Booking.end_date <= end_date)
        if status:
            query = query.filter(Booking.status == status)

        bookings = query.order_by(Booking.created_at.desc()).all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings Report"

        # Add header
        headers = [
            "Booking #", "Customer", "Phone", "Email", "Tour Rep",
            "Start Date", "End Date", "Total Amount", "Paid", "Outstanding",
            "Status", "Created At"
        ]

        # Style header
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row_idx, booking in enumerate(bookings, start=2):
            ws.cell(row=row_idx, column=1, value=booking.booking_number)
            ws.cell(row=row_idx, column=2, value=booking.customer.full_name)
            ws.cell(row=row_idx, column=3, value=booking.customer.phone or "")
            ws.cell(row=row_idx, column=4, value=booking.customer.email or "")
            ws.cell(row=row_idx, column=5, value=booking.tour_rep.full_name)
            ws.cell(row=row_idx, column=6, value=booking.start_date.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=7, value=booking.end_date.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=8, value=booking.total_amount or 0)
            ws.cell(row=row_idx, column=9, value=booking.paid_amount)
            ws.cell(row=row_idx, column=10, value=(booking.total_amount or 0) - booking.paid_amount)
            ws.cell(row=row_idx, column=11, value=booking.status.upper())
            ws.cell(row=row_idx, column=12, value=booking.created_at.strftime("%Y-%m-%d %H:%M"))

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_revenue_excel(
        self,
        db: Session,
        account_id: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> bytes:
        """Generate Excel report for revenue by tour rep"""
        # Query bookings with aggregation
        query = db.query(
            Booking.tour_rep_id,
            func.count(Booking.id).label('booking_count'),
            func.sum(Booking.total_amount).label('total_revenue'),
            func.sum(Booking.paid_amount).label('total_paid')
        ).filter(Booking.account_id == account_id)

        if start_date:
            query = query.filter(Booking.start_date >= start_date)
        if end_date:
            query = query.filter(Booking.end_date <= end_date)

        results = query.group_by(Booking.tour_rep_id).all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Revenue Report"

        # Add header
        headers = ["Tour Rep", "Bookings", "Total Revenue", "Total Paid", "Outstanding"]

        header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row_idx, result in enumerate(results, start=2):
            booking = db.query(Booking).filter(Booking.tour_rep_id == result.tour_rep_id).first()
            tour_rep_name = booking.tour_rep.full_name if booking else "Unknown"

            ws.cell(row=row_idx, column=1, value=tour_rep_name)
            ws.cell(row=row_idx, column=2, value=result.booking_count)
            ws.cell(row=row_idx, column=3, value=float(result.total_revenue or 0))
            ws.cell(row=row_idx, column=4, value=float(result.total_paid or 0))
            ws.cell(row=row_idx, column=5, value=float((result.total_revenue or 0) - (result.total_paid or 0)))

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_bookings_pdf(
        self,
        db: Session,
        account_id: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        status: Optional[str] = None
    ) -> bytes:
        """Generate PDF report for bookings"""
        # Query bookings
        query = db.query(Booking).filter(Booking.account_id == account_id)

        if start_date:
            query = query.filter(Booking.start_date >= start_date)
        if end_date:
            query = query.filter(Booking.end_date <= end_date)
        if status:
            query = query.filter(Booking.status == status)

        bookings = query.order_by(Booking.created_at.desc()).all()

        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

        # Container for elements
        elements = []

        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2563EB'),
            alignment=TA_CENTER,
            spaceAfter=30
        )
        elements.append(Paragraph("Bookings Report", title_style))
        elements.append(Spacer(1, 12))

        # Add filters info
        info_style = ParagraphStyle(
            'Info',
            parent=self.styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER
        )

        filter_info = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        if start_date:
            filter_info += f" | From: {start_date.strftime('%Y-%m-%d')}"
        if end_date:
            filter_info += f" | To: {end_date.strftime('%Y-%m-%d')}"
        if status:
            filter_info += f" | Status: {status.upper()}"

        elements.append(Paragraph(filter_info, info_style))
        elements.append(Spacer(1, 20))

        # Create table
        data = [["Booking #", "Customer", "Tour Rep", "Dates", "Amount", "Status"]]

        for booking in bookings:
            data.append([
                booking.booking_number,
                booking.customer.full_name,
                booking.tour_rep.full_name,
                f"{booking.start_date.strftime('%Y-%m-%d')}\nto\n{booking.end_date.strftime('%Y-%m-%d')}",
                f"LKR {booking.total_amount:,.2f}" if booking.total_amount else "N/A",
                booking.status.upper()
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)

        # Add summary
        elements.append(Spacer(1, 30))
        summary_style = ParagraphStyle(
            'Summary',
            parent=self.styles['Normal'],
            fontSize=12,
            alignment=TA_RIGHT
        )

        total_bookings = len(bookings)
        total_amount = sum(b.total_amount or 0 for b in bookings)
        total_paid = sum(b.paid_amount for b in bookings)
        total_outstanding = total_amount - total_paid

        summary_text = f"""
        <b>Total Bookings:</b> {total_bookings}<br/>
        <b>Total Amount:</b> LKR {total_amount:,.2f}<br/>
        <b>Total Paid:</b> LKR {total_paid:,.2f}<br/>
        <b>Total Outstanding:</b> LKR {total_outstanding:,.2f}
        """

        elements.append(Paragraph(summary_text, summary_style))

        # Build PDF
        doc.build(elements)

        buffer.seek(0)
        return buffer.getvalue()

    def generate_payments_excel(
        self,
        db: Session,
        account_id: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> bytes:
        """Generate Excel report for payments"""
        # Query payments
        query = db.query(Payment).filter(Payment.account_id == account_id)

        if start_date:
            query = query.filter(Payment.payment_date >= start_date)
        if end_date:
            query = query.filter(Payment.payment_date <= end_date)

        payments = query.order_by(Payment.payment_date.desc()).all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments Report"

        # Add header
        headers = [
            "Payment Date", "Booking #", "Customer", "Amount", "Method",
            "Status", "Receipt #", "Recorded At"
        ]

        header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row_idx, payment in enumerate(payments, start=2):
            booking = db.query(Booking).filter(Booking.id == payment.booking_id).first()

            ws.cell(row=row_idx, column=1, value=payment.payment_date.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=2, value=booking.booking_number if booking else "N/A")
            ws.cell(row=row_idx, column=3, value=booking.customer.full_name if booking else "N/A")
            ws.cell(row=row_idx, column=4, value=payment.amount)
            ws.cell(row=row_idx, column=5, value=payment.payment_method.upper())
            ws.cell(row=row_idx, column=6, value=payment.payment_status.upper())
            ws.cell(row=row_idx, column=7, value=payment.receipt_number or "")
            ws.cell(row=row_idx, column=8, value=payment.created_at.strftime("%Y-%m-%d %H:%M"))

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# Global instance
report_service = ReportService()
